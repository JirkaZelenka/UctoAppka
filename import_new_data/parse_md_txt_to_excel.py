"""
Načte .md nebo .txt s daty na vlastním řádku a položkami pod ním, vyexportuje Excel.

Rok u řádku bez roku: výchozí pro první hlavičku dává --start-rok (default 2024),
u dalších bezeslovných se předpokládá chronologické pořadí sekcí v souboru.

Navíc: hlavička může být i „ne – po 9.9.“ (datum nakonci). Řádky bez částky
(kromě dočasného technického „šumu“) jdou do Excelu jen s `radek`, `datum` a
`typ` = `?` — ostatní sloupce zůstávají prázdné.

Druhý soubor: vedle vystup.xlsx uloží ``vystup_denni.xlsx`` (jeden řádek / den, časy H:MM,
sloupec text, pokud je u daného dne uveden bezeslový blok před první účtovou položkou).

Spuštění:
  python parse_md_txt_to_excel.py vstup.txt vystup.xlsx
  python parse_md_txt_to_excel.py vstup.md
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd

def _je_radek_komentar_mrizka(t: str) -> bool:
    """# komentář (jedno #), ne ## nadpis z markdownu."""
    if t.startswith("##"):
        return False
    return t.startswith("#")


# Řádek samotné datum, např. 14.2. nebo 14.2. 2026
DATE_LINE = re.compile(
    r"^\s*(\d{1,2})\.(\d{1,2})\.(?:\s*(\d{4})\s*)?$"
)
# Datum (den.měsíc.) nakonci řádku — může předcházet třebas „ne – po 9.9.“
_DATE_NA_KONCI = re.compile(
    r"(\d{1,2})\.(\d{1,2})\.\s*(\d{4})?\s*$"
)
# Řádek začínající celým číslem v korunách: volitelné +/−, pak číslice bez mezer a desetin
AMOUNT_LINE = re.compile(r"^\s*([+-]?)\s*(\d+)\s*(.*)$")

# „Šum“: ignorované ne-záznamové řádky (komentáře mimo # řeší jinde)
_MDC_HLAVICKA = re.compile(r"^#{1,6}\s+\S")
_CARA = re.compile(r"^[\s\-\*_=·]{3,}\s*$")

# První řádek s datem *bez* roku: platí se jako rok; další bezeslovné datumy
# se doplňují chronologicky (pokud m/d jde v kalendáři „dozadu“ oproti předchozímu, přičte se +1 k roku).
# Nejednoznačné věci (dva kratší bloky stejného dne ve dvou různých letech) raději s rokem v textu.
DEFAULT_START_YEAR = 2024

# Pořadí sloupců v Excelu: původní řádek, datum, dál částka a doplnky
SLOUPCE = (
    "radek",
    "datum",
    "typ",
    "castka",
    "kontrola_nasobek_5",
    "text",
    "spolecny",
    "měsíce",
)

# Zjednodušený export „deník“: jeden řádek / den, časy H:MM, text
SLOUPCE_DENNI = ("radek", "datum", "cas1", "cas2", "cas3", "text")

# Čas H:MM nebo H:M v řádku (1:0 i 2:15, max. rozumné číslo hodin u záznamu spánku)
_CAS_HMM = re.compile(r"\b(\d{1,2}):(\d{1,2})\b")

# Text končí „N měsíc“ / „N měsíce“ / „N měsíců“ (case insensitive)
_MESICE_KONEC = re.compile(r"(?i)(\d+)\s+měsíc(?:e|ů)?\s*$")

# Jednoduchá normalizace diakritiky pro detekci „společný“
_DIACRITIC_FROM = "áčďéěíňóřšťúůýžÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ"
_DIACRITIC_TO = "acdeeinorstuuyzACDEEINORSTUUYZ"
_TRANSL = str.maketrans(_DIACRITIC_FROM, _DIACRITIC_TO)


def _normalize_word(s: str) -> str:
    return s.translate(_TRANSL).lower()


def _is_spolecny_text(rest: str) -> bool:
    if not rest or not rest.strip():
        return False
    first = rest.strip().split(maxsplit=1)[0]
    return _normalize_word(first).startswith("spolec")


def _mesice_z_textu(text: str) -> int | None:
    s = text.strip()
    if not s:
        return None
    m = _MESICE_KONEC.search(s)
    if not m:
        return None
    return int(m.group(1))


def _parse_amount(sign: str, num_raw: str) -> int | None:
    if not num_raw.isdigit():
        return None
    v = int(num_raw)
    if sign == "-":
        v = -v
    elif sign == "+":
        v = abs(v)
    return v


def _classify_line_kind_and_amount(line: str) -> tuple[str, int | None, str]:
    """
    Vrátí (typ, částka|None, zbytek řádku).
    typ: 'výdaj' | 'příjem' | '?'
    """
    m = AMOUNT_LINE.match(line)
    if not m:
        return "?", None, line.strip()

    sign, num_raw, rest = m.group(1), m.group(2), m.group(3).strip()

    amount = _parse_amount(sign, num_raw)
    if amount is None:
        return "?", None, line.strip()

    if amount < 0:
        kind = "výdaj"
    else:
        kind = "příjem"
    return kind, amount, rest


def _ma_radek_castku(stripped: str) -> bool:
    m = AMOUNT_LINE.match(stripped)
    if not m:
        return False
    if _parse_amount(m.group(1), m.group(2)) is None:
        return False
    zbytek = (m.group(3) or "").lstrip()
    if zbytek.startswith(":") and re.match(
        r"^:([0-5]?\d)\b", zbytek
    ):
        # „2:15“ není 2 Kč, ale čas
        return False
    return True


def _parsovat_hlavicku_datum(stripped: str) -> tuple[int, int, int | None] | None:
    """
    (den, měsíc, rok|None) pro řádek s datem, nebo None.
    Povolí i „ne – po 9.9.“; když řádek vypadá jako transakce (začíná částkou), bere se jako vklad.
    """
    m0 = DATE_LINE.match(stripped)
    if m0:
        d, mo, yg = int(m0.group(1)), int(m0.group(2)), m0.group(3)
        return (d, mo, int(yg) if yg else None)
    m = _DATE_NA_KONCI.search(stripped)
    if not m:
        return None
    if m.start() > 0:
        pred = stripped[: m.start()]
        if pred.strip() and not re.fullmatch(
            r"[\s\-–—,;:A-Za-zÁ-ž0-9()&%'+\"„“]+",
            pred,
        ):
            return None
    if _ma_radek_castku(stripped):
        return None
    d, mo = int(m.group(1)), int(m.group(2))
    yg = m.group(3)
    return (d, mo, int(yg) if yg else None)


def _je_hluk(stripped: str) -> bool:
    t = stripped.strip()
    if t.startswith("http://") or t.startswith("https://") or t.startswith("www."):
        return True
    if t.startswith(("@", "mailto:")):
        return True
    if t.startswith(">"):
        if _MDC_HLAVICKA.match(t) or t.startswith("> "):
            return True
    if t.startswith("```") or t.startswith("~~~"):
        return True
    if t.startswith(("<div", "<span", "<!--")):
        return True
    if _MDC_HLAVICKA.match(t):
        return True
    if _CARA.match(t) or t in ("—", "–", "-", "*", "___", "---", "==="):
        return True
    if re.match(r"^!?\[[^]]*\]\([^)]+\)\s*$", t):
        return True
    return False


def _dopocitat_datum_bez_roku(
    m: int, d: int, posledni: date | None, start_rok: int
) -> date:
    y = start_rok if posledni is None else posledni.year
    for _ in range(80):
        try:
            cand = date(y, m, d)
        except ValueError:
            y += 1
            continue
        if posledni is None or cand >= posledni:
            return cand
        y += 1
    raise ValueError("Nepodařilo se dovodit rok (zkontroluj pořadí sekcí nebo uveď rok v textu).")


def _append_radek_bez_s_castky(
    out: list[dict],
    line: str,
    dtm: date | None,
) -> None:
    """Řádek bez rozpoznané částky: jen 1) původní text, 2) datum, 3) typ „?“."""
    out.append(
        {
            "radek": line,
            "datum": dtm,
            "typ": "?",
            "castka": None,
            "kontrola_nasobek_5": "",
            "text": "",
            "spolecny": "",
            "měsíce": None,
        }
    )


def _append_row(
    out: list[dict],
    line: str,
    dtm: date | None,
    kind: str,
    amount_signed: int | None,
    zbytek: str,
) -> None:
    castka = abs(amount_signed) if amount_signed is not None else None
    kontrola = ""
    if castka is not None and castka % 5 != 0:
        kontrola = "?"
    spolecny = "S" if _is_spolecny_text(zbytek) else ""
    mesice = _mesice_z_textu(zbytek)
    out.append(
        {
            "radek": line,
            "datum": dtm,
            "typ": kind,
            "castka": castka,
            "kontrola_nasobek_5": kontrola,
            "text": zbytek,
            "spolecny": spolecny,
            "měsíce": mesice,
        }
    )


def parse_lines(lines: list[str], start_rok: int = DEFAULT_START_YEAR) -> list[dict]:
    bloky: list[tuple[str, str]] = []
    for raw in lines:
        line = raw.rstrip("\n\r")
        t = line.strip()
        if not t or _je_radek_komentar_mrizka(t):
            continue
        bloky.append((line, t))

    rows: list[dict] = []
    current_date: date | None = None
    posledni_hlavicka: date | None = None

    for line, stripped in bloky:
        h = _parsovat_hlavicku_datum(stripped)
        if h is not None:
            d, mo, yg = h
            if yg is not None:
                current_date = date(yg, mo, d)
            else:
                current_date = _dopocitat_datum_bez_roku(mo, d, posledni_hlavicka, start_rok)
            posledni_hlavicka = current_date
            continue

        if current_date is None:
            if _je_hluk(stripped):
                continue
            _append_radek_bez_s_castky(rows, line, None)
            continue

        if _je_hluk(stripped):
            continue

        if _ma_radek_castku(stripped):
            kind, amount_signed, rest = _classify_line_kind_and_amount(stripped)
            _append_row(rows, line, current_date, kind, amount_signed, rest)
        else:
            _append_radek_bez_s_castky(rows, line, current_date)

    return rows


def _denni_tii_casy_a_text(slines: list[str]) -> tuple[str | None, str | None, str | None, str]:
    """
    První řádek s H:MM. Jen pokud jsou na tom řádku přesně 3 časy, vyplní cas1–3;
    jinak tři sloupce „?“ (doplní se ručně). Text = zbytek deníku v logickém pořadí.
    """
    t_idx = -1
    mlist_line: list[re.Match[str]] = []
    for i, s in enumerate(slines):
        mlist = list(_CAS_HMM.finditer(s))
        if mlist:
            t_idx = i
            mlist_line = mlist
            break
    if t_idx < 0:
        flat = " ".join(x.strip() for x in slines if x.strip())
        return (None, None, None, flat)
    tline = slines[t_idx]
    mlist = mlist_line
    n_all = len(mlist)
    if n_all == 3:
        t1 = f"{mlist[0].group(1)}:{mlist[0].group(2)}"
        t2 = f"{mlist[1].group(1)}:{mlist[1].group(2)}"
        t3 = f"{mlist[2].group(1)}:{mlist[2].group(2)}"
    else:
        t1 = t2 = t3 = "?"
    n_tail = n_all if n_all else 0
    tail0 = (
        tline[mlist[n_tail - 1].end() :].lstrip(" \t,;:").strip()
        if mlist and n_tail
        else ""
    )
    parts: list[str] = []
    for s in slines[:t_idx]:
        if s.strip():
            parts.append(s.strip())
    if tail0:
        parts.append(tail0)
    for s in slines[t_idx + 1 :]:
        if s.strip():
            parts.append(s.strip())
    return (t1, t2, t3, " ".join(parts).strip())


def parse_denni_lines(lines: list[str], start_rok: int = DEFAULT_START_YEAR) -> list[dict]:
    """
    Jednořádkový denní přehled: pod každým datem (hlavičkou) sbírá neúčetní řádky,
    dokud nenarazí na řádek s částkou (účto) nebo na další datum. Časy hledá
    v prvním řádku, kde se vyskytne H:MM; více řádků spojí.
    Na tom řádku musí být přesně tři časy, jinak jsou v Excelu v cas1–3 otazníky.
    """
    bloky: list[tuple[str, str]] = []
    for raw in lines:
        line = raw.rstrip("\n\r")
        t = line.strip()
        if not t or _je_radek_komentar_mrizka(t):
            continue
        bloky.append((line, t))

    rows: list[dict] = []
    current_date: date | None = None
    posledni_hlavicka: date | None = None
    buf: list[tuple[str, str]] = []

    def _flush() -> None:
        nonlocal buf, current_date
        if not buf or current_date is None:
            return
        strips_only = [b for _, b in buf]
        t1, t2, t3, txt = _denni_tii_casy_a_text(strips_only)
        radek = " ".join(a.strip() for a, _ in buf)
        rows.append(
            {
                "radek": radek,
                "datum": current_date,
                "cas1": t1,
                "cas2": t2,
                "cas3": t3,
                "text": txt,
            }
        )
        buf = []

    for line, stripped in bloky:
        h = _parsovat_hlavicku_datum(stripped)
        if h is not None:
            _flush()
            d, mo, yg = h
            if yg is not None:
                current_date = date(yg, mo, d)
            else:
                current_date = _dopocitat_datum_bez_roku(
                    mo, d, posledni_hlavicka, start_rok
                )
            posledni_hlavicka = current_date
            continue

        if current_date is None:
            continue
        if _ma_radek_castku(stripped):
            _flush()
            continue
        if _je_hluk(stripped):
            continue
        buf.append((line, stripped))
    _flush()
    return rows


def _excel_bunky_cele_cisla(path: Path, sloupce: tuple[str, ...]) -> None:
    """Po uložení přes pandas nastaví u vybraných sloupců hodnotu jako int (ne float)."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    hlavicky = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    }
    for jmeno in sloupce:
        col = hlavicky.get(jmeno)
        if col is None:
            continue
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col)
            v = cell.value
            if v is None or v == "":
                continue
            try:
                cell.value = int(v)
            except (TypeError, ValueError):
                continue
            cell.number_format = "0"
    wb.save(path)


def parse_file(path: Path, start_rok: int = DEFAULT_START_YEAR) -> list[dict]:
    text = path.read_text(encoding="utf-8")
    lines = text.splitlines()
    return parse_lines(lines, start_rok=start_rok)


def parse_denni_file(
    path: Path, start_rok: int = DEFAULT_START_YEAR
) -> list[dict]:
    text = path.read_text(encoding="utf-8")
    lines = text.splitlines()
    return parse_denni_lines(lines, start_rok=start_rok)


def main() -> int:
    p = argparse.ArgumentParser(description="Parsuje .md/.txt do Excelu.")
    p.add_argument("vstup", type=Path, help="Cesta k .txt nebo .md")
    p.add_argument(
        "vystup",
        type=Path,
        nargs="?",
        help="Cesta k .xlsx (výchozí: stejný název jako vstup + .xlsx). Uloží se i *vstup_denni.xlsx*.",
    )
    p.add_argument(
        "--start-rok",
        type=int,
        default=DEFAULT_START_YEAR,
        dest="start_rok",
        metavar="ROK",
        help=(
            f"Rok u prvního data *bez* roku v souboru (výchozí {DEFAULT_START_YEAR}); "
            "další bezeslovná se odvíjejí chronologicky oproti předchozímu."
        ),
    )
    p.add_argument(
        "--rok",
        type=int,
        default=None,
        help=f"Zastaralé, stejné jako --start-rok. Pokud zadáš, přebije --start-rok.",
    )
    args = p.parse_args()
    start = args.rok if args.rok is not None else args.start_rok

    if not args.vstup.is_file():
        print(f"Soubor neexistuje: {args.vstup}", file=sys.stderr)
        return 1

    out = args.vystup or args.vstup.with_suffix(".xlsx")
    rows = parse_file(args.vstup, start_rok=start)

    df = pd.DataFrame(rows, columns=list(SLOUPCE))
    # Excel datum jako datum
    if "datum" in df.columns and len(df):
        df["datum"] = pd.to_datetime(df["datum"], errors="coerce").dt.date

    if "castka" in df.columns:
        df["castka"] = df["castka"].astype("Int64")
    if "měsíce" in df.columns:
        df["měsíce"] = df["měsíce"].astype("Int64")

    df.to_excel(out, index=False, engine="openpyxl")
    _excel_bunky_cele_cisla(out, ("castka", "měsíce"))
    print(f"Zapsáno: {out} ({len(df)} řádků)")

    denni_out = out.parent / f"{out.stem}_denni{out.suffix}"
    rows_d = parse_denni_file(args.vstup, start_rok=start)
    dfd = pd.DataFrame(rows_d, columns=list(SLOUPCE_DENNI))
    if "datum" in dfd.columns and len(dfd):
        dfd["datum"] = pd.to_datetime(dfd["datum"], errors="coerce").dt.date
    dfd.to_excel(denni_out, index=False, engine="openpyxl")
    print(f"Zapsáno (denní přehled): {denni_out} ({len(dfd)} řádků)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
