import os
import subprocess

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Q, Count, Avg, Min, Max
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from datetime import datetime, timedelta
from decimal import Decimal
import statistics
import plotly.graph_objects as go
import plotly.express as px
from plotly.offline import plot
import csv
import io
try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    Transaction, Category, Subcategory, RecurringPayment,
    Investment, InvestmentObservation, BudgetLimit, TransactionType, PaymentFor
)
from .forms import TransactionForm, CategoryForm, SubcategoryForm, InvestmentForm, InvestmentValueForm, RecurringPaymentForm


def _first_env_value(names):
    for name in names:
        value = (os.environ.get(name) or "").strip()
        if value:
            return value
    return ""


def _git_short_commit():
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
            timeout=1,
        )
        return (result.stdout or "").strip()
    except Exception:
        return ""


def _git_commit_date_iso():
    try:
        result = subprocess.run(
            ["git", "show", "-s", "--format=%cI", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
            timeout=1,
        )
        return (result.stdout or "").strip()
    except Exception:
        return ""


@require_http_methods(["GET"])
def health(_request):
    # Prefer deployment-provided values, then fall back to local git metadata.
    version = _first_env_value(
        [
            "APP_VERSION",
            "RELEASE_VERSION",
            "FLY_IMAGE_REF",
            "FLY_RELEASE_ID",
            "GITHUB_REF_NAME",
            "GITHUB_RUN_NUMBER",
        ]
    )
    commit = _first_env_value(
        [
            "APP_COMMIT",
            "COMMIT_SHA",
            "GIT_COMMIT",
            "SOURCE_COMMIT",
            "GITHUB_SHA",
        ]
    )
    if not commit:
        commit = _git_short_commit()
    commit_date = _first_env_value(["APP_COMMIT_DATE", "COMMIT_DATE", "GITHUB_EVENT_HEAD_COMMIT_TIMESTAMP"])
    if not commit_date:
        commit_date = _git_commit_date_iso()

    fly_info = {
        "app": _first_env_value(["FLY_APP_NAME"]),
        "region": _first_env_value(["FLY_REGION"]),
        "machine_id": _first_env_value(["FLY_MACHINE_ID"]),
        "instance_id": _first_env_value(["FLY_ALLOC_ID"]),
        "release_id": _first_env_value(["FLY_RELEASE_ID"]),
        "image_ref": _first_env_value(["FLY_IMAGE_REF"]),
    }

    return JsonResponse(
        {
            "status": "ok",
            "timestamp": timezone.now().isoformat(),
            "version": version,
            "commit": commit,
            "commit_date": commit_date,
            "fly": fly_info,
        }
    )


def calculate_split_amounts(transactions):
    """Calculate amounts split between user 1 (SELF) and user 2 (PARTNER).
    SHARED transactions are split 50/50.
    Returns: (user1_amount, user2_amount)
    """
    user1_total = Decimal('0')
    user2_total = Decimal('0')
    
    for transaction in transactions:
        amount = transaction.amount
        if transaction.payment_for == PaymentFor.SELF:
            user1_total += amount
        elif transaction.payment_for == PaymentFor.PARTNER:
            user2_total += amount
        elif transaction.payment_for == PaymentFor.SHARED:
            # Split 50/50
            half = amount / 2
            user1_total += half
            user2_total += half
    
    return user1_total, user2_total


def get_split_user_labels():
    """Display labels for the SELF/PARTNER split cards."""
    preferred_order = ['jirka', 'zuzka']
    usernames = list(
        User.objects.filter(username__in=preferred_order)
        .order_by('id')
        .values_list('username', flat=True)
    )
    ordered = [name for name in preferred_order if name in usernames]
    if len(ordered) < 2:
        fallback = list(
            User.objects.exclude(username='admin')
            .order_by('id')
            .values_list('username', flat=True)[:2]
        )
        ordered = fallback if len(fallback) == 2 else ['jirka', 'zuzka']
    return ordered[0], ordered[1]


@login_required
def dashboard(request):
    """Měsíční dashboard"""
    # Toggle mezi kalendářním měsícem a posledními 30 dny
    period = request.GET.get('period', 'month')  # 'month' or '30days'
    
    today = timezone.now().date()
    
    if period == '30days':
        start_date = today - timedelta(days=30)
        end_date = today
        period_label = "Posledních 30 dní"
    else:
        start_date = today.replace(day=1)
        end_date = today
        period_label = f"{today.strftime('%B %Y')}"
    
    # Filtrování transakcí (exclude deleted)
    transactions = Transaction.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        is_deleted=False
    )
    
    # Filtry
    transaction_type_filter = request.GET.get('type', '')
    if transaction_type_filter:
        transactions = transactions.filter(transaction_type=transaction_type_filter)
    
    category_filter = request.GET.get('category', '')
    if category_filter:
        transactions = transactions.filter(category_id=category_filter)
    
    # Řazení
    sort_by = request.GET.get('sort', 'date')
    sort_order = request.GET.get('order', 'desc')
    
    # Mapování názvů sloupců na pole modelu
    sort_fields = {
        'date': 'date',
        'description': 'description',
        'type': 'transaction_type',
        'category': 'category__name',
        'amount': 'amount',
        'created_by': 'created_by__username',
        'created_at': 'created_at',
        'payment_for': 'payment_for',
        'approved': 'approved',
    }
    
    # Výchozí řazení
    if sort_by not in sort_fields:
        sort_by = 'date'
    
    order_field = sort_fields[sort_by]
    
    # Aplikovat řazení
    if sort_order == 'asc':
        transactions = transactions.order_by(order_field, '-created_at')
    else:
        transactions = transactions.order_by(f'-{order_field}', '-created_at')
    
    # Statistiky (před limitováním)
    income_transactions = transactions.filter(transaction_type=TransactionType.INCOME)
    expense_transactions = transactions.filter(transaction_type=TransactionType.EXPENSE)
    investment_transactions = transactions.filter(transaction_type=TransactionType.INVESTMENT)
    
    income = income_transactions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    expenses = expense_transactions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    # Přesuny (investice) - součet transakcí typu INVESTMENT
    investments = investment_transactions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    net_income = income - expenses
    
    # Calculate split amounts
    income_user1, income_user2 = calculate_split_amounts(income_transactions)
    expenses_user1, expenses_user2 = calculate_split_amounts(expense_transactions)
    investments_user1, investments_user2 = calculate_split_amounts(investment_transactions)
    net_income_user1 = income_user1 - expenses_user1
    net_income_user2 = income_user2 - expenses_user2
    
    # Poslední transakce (limit 20)
    recent_transactions = transactions[:20]
    
    categories = Category.objects.all()
    
    split_user1_label, split_user2_label = get_split_user_labels()

    context = {
        'transactions': recent_transactions,
        'income': income,
        'expenses': expenses,
        'investments': investments,
        'net_income': net_income,
        'income_user1': income_user1,
        'income_user2': income_user2,
        'expenses_user1': expenses_user1,
        'expenses_user2': expenses_user2,
        'investments_user1': investments_user1,
        'investments_user2': investments_user2,
        'net_income_user1': net_income_user1,
        'net_income_user2': net_income_user2,
        'period': period,
        'period_label': period_label,
        'categories': categories,
        'transaction_type_filter': transaction_type_filter,
        'category_filter': category_filter,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'split_user1_label': split_user1_label,
        'split_user2_label': split_user2_label,
    }
    
    return render(request, 'expenses/dashboard.html', context)


@login_required
def manage_transactions(request):
    """Správa transakcí - Přidat, Import/Export, Spravuj"""
    # Handle adding new transaction
    if request.method == 'POST' and 'add_transaction' in request.POST:
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created_by = request.user
            transaction.approved = False
            transaction.save()
            messages.success(request, 'Transakce byla úspěšně přidána.')
            return redirect(reverse('manage_transactions') + '?tab=manage')
    else:
        form = TransactionForm(
            initial={
                'created_by': request.user,
                'date': timezone.now().date().strftime('%Y-%m-%d'),
            }
        )
    
    # Get all transactions for "Spravuj transakce" section (exclude deleted)
    transactions = Transaction.objects.filter(is_deleted=False)
    
    # Get min and max dates from all transactions for default filter values
    date_range = Transaction.objects.aggregate(
        min_date=Min('date'),
        max_date=Max('date')
    )
    
    # Filters
    transaction_type_filter = request.GET.get('type', '')
    if transaction_type_filter:
        transactions = transactions.filter(transaction_type=transaction_type_filter)
    
    category_filter = request.GET.get('category', '')
    if category_filter:
        transactions = transactions.filter(category_id=category_filter)
    
    # Date filters - use GET params if provided, otherwise use defaults
    date_from = request.GET.get('date_from', '')
    if not date_from and date_range['min_date']:
        date_from = date_range['min_date'].strftime('%Y-%m-%d')
    
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if not date_to and date_range['max_date']:
        date_to = date_range['max_date'].strftime('%Y-%m-%d')
    
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    approved_filter = request.GET.get('approved', '')
    if approved_filter == 'yes':
        transactions = transactions.filter(approved=True)
    elif approved_filter == 'no':
        transactions = transactions.filter(approved=False)
    
    # Sorting
    sort_by = request.GET.get('sort', 'date')
    sort_order = request.GET.get('order', 'desc')
    
    sort_fields = {
        'date': 'date',
        'description': 'description',
        'type': 'transaction_type',
        'category': 'category__name',
        'amount': 'amount',
        'created_by': 'created_by__username',
        'created_at': 'created_at',
        'payment_for': 'payment_for',
        'approved': 'approved',
    }
    
    if sort_by not in sort_fields:
        sort_by = 'date'
    
    order_field = sort_fields[sort_by]
    
    if sort_order == 'asc':
        transactions = transactions.order_by(order_field, '-created_at')
    else:
        transactions = transactions.order_by(f'-{order_field}', '-created_at')
    
    categories = Category.objects.all()
    
    # Get min and max dates for export form
    date_range = Transaction.objects.filter(is_deleted=False).aggregate(
        min_date=Min('date'),
        max_date=Max('date')
    )
    
    context = {
        'form': form,
        'transactions': transactions,
        'categories': categories,
        'transaction_type_filter': transaction_type_filter,
        'category_filter': category_filter,
        'date_from': date_from,
        'date_to': date_to,
        'approved_filter': approved_filter,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'date_range': date_range,
    }
    
    return render(request, 'expenses/manage_transactions.html', context)


@login_required
def add_transaction(request):
    """Přidání nové transakce - redirect to manage_transactions"""
    return redirect('manage_transactions')


@login_required
def edit_transaction(request, pk):
    """Editace transakce"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transakce byla úspěšně upravena.')
            # Redirect back to the page that called this (dashboard or manage_transactions)
            referer = request.META.get('HTTP_REFERER', '')
            if 'manage-transactions' in referer:
                return redirect(reverse('manage_transactions') + '?tab=manage')
            else:
                return redirect('dashboard')
    else:
        # Prefill form with transaction data, ensuring date is set
        form = TransactionForm(instance=transaction)
        # Ensure date is properly formatted for the date input (YYYY-MM-DD format)
        if transaction.date:
            form.initial['date'] = transaction.date.strftime('%Y-%m-%d')
    
    return render(request, 'expenses/edit_transaction.html', {'form': form, 'transaction': transaction})


@login_required
def approve_transaction(request, pk):
    """Schválení transakce"""
    transaction = get_object_or_404(Transaction, pk=pk)
    transaction.approved = True
    transaction.save()
    messages.success(request, 'Transakce byla schválena.')
    # Redirect back to the page that called this (dashboard or manage_transactions)
    referer = request.META.get('HTTP_REFERER', '')
    if 'manage-transactions' in referer:
        return redirect(reverse('manage_transactions') + '?tab=manage')
    else:
        return redirect('dashboard')


@login_required
def statistics(request):
    """Dlouhodobé statistiky s grafy"""
    # Filtry
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    category_filter = request.GET.get('category', '')
    
    transactions = Transaction.objects.filter(is_deleted=False)
    
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    if category_filter:
        transactions = transactions.filter(category_id=category_filter)
    
    # Měsíční statistiky
    monthly_stats = transactions.values('date__year', 'date__month').annotate(
        income=Sum('amount', filter=Q(transaction_type=TransactionType.INCOME)),
        expenses=Sum('amount', filter=Q(transaction_type=TransactionType.EXPENSE)),
        investments=Sum('amount', filter=Q(transaction_type=TransactionType.INVESTMENT)),
    ).order_by('date__year', 'date__month')
    
    # Příprava dat pro graf
    months = []
    incomes = []
    expenses_list = []
    net_incomes = []
    investments_list = []
    
    for stat in monthly_stats:
        month_str = f"{stat['date__year']}-{stat['date__month']:02d}"
        months.append(month_str)
        income = stat['income'] or Decimal('0')
        expense = stat['expenses'] or Decimal('0')
        investment = stat['investments'] or Decimal('0')
        incomes.append(float(income))
        expenses_list.append(float(expense))
        net_incomes.append(float(income - expense))
        investments_list.append(float(investment))
    
    # Graf měsíčních statistik
    fig_monthly = go.Figure()
    if months:  # Only add traces if there's data
        fig_monthly.add_trace(go.Scatter(x=months, y=incomes, name='Příjmy', mode='lines+markers', line=dict(color='green')))
        fig_monthly.add_trace(go.Scatter(x=months, y=expenses_list, name='Výdaje', mode='lines+markers', line=dict(color='red')))
        fig_monthly.add_trace(go.Scatter(x=months, y=net_incomes, name='Čistý příjem', mode='lines+markers', line=dict(color='blue')))
        fig_monthly.add_trace(go.Scatter(x=months, y=investments_list, name='Investice', mode='lines+markers', line=dict(color='orange')))
    fig_monthly.update_layout(
        title='Měsíční přehled - Příjmy, Výdaje, Čistý příjem, Investice',
        xaxis_title='Měsíc',
        yaxis_title='Částka (Kč)',
        hovermode='x unified'
    )
    plot_monthly = plot(fig_monthly, output_type='div', include_plotlyjs='cdn')
    
    # Koláčový graf - kategorie
    view_type = request.GET.get('view', 'category')  # 'category' or 'subcategory'
    
    if view_type == 'subcategory':
        category_stats = transactions.filter(transaction_type=TransactionType.EXPENSE).values(
            'subcategory__name'
        ).annotate(total=Sum('amount')).order_by('-total')[:10]
        labels = [s['subcategory__name'] or 'Bez subkategorie' for s in category_stats]
    else:
        category_stats = transactions.filter(transaction_type=TransactionType.EXPENSE).values(
            'category__name'
        ).annotate(total=Sum('amount')).order_by('-total')[:10]
        labels = [s['category__name'] or 'Bez kategorie' for s in category_stats]
    
    values = [float(s['total']) for s in category_stats]
    
    if labels and values:  # Only create pie chart if there's data
        fig_pie = go.Figure(data=[go.Pie(labels=labels, values=values)])
        fig_pie.update_layout(title='Rozpad výdajů podle kategorií' if view_type == 'category' else 'Rozpad výdajů podle subkategorií')
    else:
        fig_pie = go.Figure()
        fig_pie.add_annotation(text='Žádná data k zobrazení', xref='paper', yref='paper', x=0.5, y=0.5, showarrow=False)
        fig_pie.update_layout(title='Rozpad výdajů podle kategorií' if view_type == 'category' else 'Rozpad výdajů podle subkategorií')
    plot_pie = plot(fig_pie, output_type='div', include_plotlyjs='cdn')
    
    # Celkové čisté jmění
    total_income = transactions.filter(transaction_type=TransactionType.INCOME).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    total_expenses = transactions.filter(transaction_type=TransactionType.EXPENSE).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    total_investments = transactions.filter(transaction_type=TransactionType.INVESTMENT).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    # Aktuální hodnota investic
    current_investment_value = Investment.objects.aggregate(
        total=Sum('observed_value')
    )['total'] or Decimal('0')
    
    net_worth = total_income - total_expenses - total_investments + current_investment_value
    
    # Graf čistého jmění v čase
    cumulative_net = []
    running_total = Decimal('0')
    for stat in monthly_stats:
        income = stat['income'] or Decimal('0')
        expense = stat['expenses'] or Decimal('0')
        investment = stat['investments'] or Decimal('0')
        running_total += income - expense - investment
        cumulative_net.append(float(running_total))
    
    fig_networth = go.Figure()
    if months and cumulative_net:  # Only add trace if there's data
        fig_networth.add_trace(go.Scatter(x=months, y=cumulative_net, name='Čisté jmění', mode='lines+markers', line=dict(color='purple')))
    fig_networth.update_layout(
        title='Vývoj čistého jmění v čase',
        xaxis_title='Měsíc',
        yaxis_title='Částka (Kč)',
        hovermode='x unified'
    )
    plot_networth = plot(fig_networth, output_type='div', include_plotlyjs='cdn')
    
    categories = Category.objects.all()
    
    context = {
        'plot_monthly': plot_monthly,
        'plot_pie': plot_pie,
        'plot_networth': plot_networth,
        'categories': categories,
        'view_type': view_type,
        'start_date': start_date,
        'end_date': end_date,
        'category_filter': category_filter,
        'net_worth': net_worth,
    }
    
    return render(request, 'expenses/statistics.html', context)


@login_required
def predictions(request):
    """Predikce a očekávané výdaje - aktuální měsíc"""
    today = timezone.now().date()
    current_month_start = today.replace(day=1)
    
    # Vypočítat konec aktuálního měsíce
    from calendar import monthrange
    last_day = monthrange(today.year, today.month)[1]
    current_month_end = today.replace(day=last_day)
    
    # Očekávané hodnoty z trvalých plateb pro aktuální měsíc
    recurring_payments = RecurringPayment.objects.filter(active=True)
    expected_income = Decimal('0')
    expected_expenses = Decimal('0')
    expected_income_user1 = Decimal('0')
    expected_income_user2 = Decimal('0')
    expected_expenses_user1 = Decimal('0')
    expected_expenses_user2 = Decimal('0')
    expected_recurring_list = []
    
    for rp in recurring_payments:
        # Použít typ transakce přímo z trvalé platby
        transaction_type = rp.transaction_type
        
        # Zkontrolovat, zda by tato trvalá platba měla proběhnout v aktuálním měsíci
        # Zkontrolovat, zda next_payment_date je v aktuálním měsíci
        # NEBO zda už proběhla v aktuálním měsíci (existuje transakce s datem v aktuálním měsíci)
        payment_date = rp.next_payment_date
        should_occur_this_month = current_month_start <= payment_date <= current_month_end
        
        # Pokud next_payment_date není v aktuálním měsíci, zkontrolovat, zda už proběhla
        if not should_occur_this_month:
            # Zkontrolovat, zda existuje transakce pro tuto trvalou platbu v aktuálním měsíci
            has_transaction_this_month = Transaction.objects.filter(
                recurring_payment=rp,
                date__gte=current_month_start,
                date__lte=current_month_end
            ).exists()
            
            # Nebo zkontrolovat podle detailů (kategorie, subkategorie, částka, typ)
            if not has_transaction_this_month and rp.category:
                has_transaction_this_month = Transaction.objects.filter(
                    category=rp.category,
                    subcategory=rp.subcategory,
                    amount=rp.amount,
                    transaction_type=rp.transaction_type,
                    date__gte=current_month_start,
                    date__lte=current_month_end
                ).exists()
            
            should_occur_this_month = has_transaction_this_month
        
        if should_occur_this_month:
            amount = rp.amount
            if transaction_type == TransactionType.INCOME:
                expected_income += amount
                # Split based on payment_for
                if rp.payment_for == PaymentFor.SELF:
                    expected_income_user1 += amount
                elif rp.payment_for == PaymentFor.PARTNER:
                    expected_income_user2 += amount
                elif rp.payment_for == PaymentFor.SHARED:
                    expected_income_user1 += amount / 2
                    expected_income_user2 += amount / 2
            elif transaction_type == TransactionType.EXPENSE:
                expected_expenses += amount
                # Split based on payment_for
                if rp.payment_for == PaymentFor.SELF:
                    expected_expenses_user1 += amount
                elif rp.payment_for == PaymentFor.PARTNER:
                    expected_expenses_user2 += amount
                elif rp.payment_for == PaymentFor.SHARED:
                    expected_expenses_user1 += amount / 2
                    expected_expenses_user2 += amount / 2
            
            expected_recurring_list.append({
                'payment': rp,
                'type': transaction_type,
                'amount': rp.amount
            })
    
    # Skutečné hodnoty pro aktuální měsíc - pouze pro trvalé platby
    # Najít všechny transakce, které odpovídají trvalým platbám z expected_recurring_list
    actual_income_transaction_ids = []
    actual_expense_transaction_ids = []
    
    for item in expected_recurring_list:
        rp = item['payment']
        transaction_type = item['type']
        
        # Najít transakce pro tuto trvalou platbu v aktuálním měsíci
        # 1. Transakce přímo propojené přes recurring_payment
        matching_transactions = Transaction.objects.filter(
            recurring_payment=rp,
            date__gte=current_month_start,
            date__lte=today,
            transaction_type=transaction_type
        )
        
        # 2. Transakce, které odpovídají podle detailů (kategorie, subkategorie, částka, typ)
        if rp.category:
            matching_by_details = Transaction.objects.filter(
                category=rp.category,
                subcategory=rp.subcategory,
                amount=rp.amount,
                transaction_type=transaction_type,
                date__gte=current_month_start,
                date__lte=today
            ).exclude(id__in=[t.id for t in matching_transactions])
            
            matching_transactions = list(matching_transactions) + list(matching_by_details)
        
        # Přidat do příslušného seznamu podle typu
        for trans in matching_transactions:
            if transaction_type == TransactionType.INCOME:
                if trans.id not in actual_income_transaction_ids:
                    actual_income_transaction_ids.append(trans.id)
            elif transaction_type == TransactionType.EXPENSE:
                if trans.id not in actual_expense_transaction_ids:
                    actual_expense_transaction_ids.append(trans.id)
    
    # Vytvořit querysety z nalezených transakcí
    if actual_income_transaction_ids:
        actual_income_transactions = Transaction.objects.filter(id__in=actual_income_transaction_ids)
    else:
        actual_income_transactions = Transaction.objects.none()
    
    if actual_expense_transaction_ids:
        actual_expense_transactions = Transaction.objects.filter(id__in=actual_expense_transaction_ids)
    else:
        actual_expense_transactions = Transaction.objects.none()
    
    actual_income = actual_income_transactions.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    actual_expenses = actual_expense_transactions.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Calculate split amounts for actual values
    actual_income_user1, actual_income_user2 = calculate_split_amounts(actual_income_transactions)
    actual_expenses_user1, actual_expenses_user2 = calculate_split_amounts(actual_expense_transactions)
    
    # Název aktuálního měsíce
    month_names = {
        1: 'Leden', 2: 'Únor', 3: 'Březen', 4: 'Duben',
        5: 'Květen', 6: 'Červen', 7: 'Červenec', 8: 'Srpen',
        9: 'Září', 10: 'Říjen', 11: 'Listopad', 12: 'Prosinec'
    }
    current_month_name = month_names.get(today.month, '')
    
    split_user1_label, split_user2_label = get_split_user_labels()

    context = {
        'expected_income': expected_income,
        'expected_expenses': expected_expenses,
        'expected_income_user1': expected_income_user1,
        'expected_income_user2': expected_income_user2,
        'expected_expenses_user1': expected_expenses_user1,
        'expected_expenses_user2': expected_expenses_user2,
        'actual_income': actual_income,
        'actual_expenses': actual_expenses,
        'actual_income_user1': actual_income_user1,
        'actual_income_user2': actual_income_user2,
        'actual_expenses_user1': actual_expenses_user1,
        'actual_expenses_user2': actual_expenses_user2,
        'expected_recurring_list': expected_recurring_list,
        'current_month_name': current_month_name,
        'current_year': today.year,
        'today': today,
        'current_month_end': current_month_end,
        'split_user1_label': split_user1_label,
        'split_user2_label': split_user2_label,
    }
    
    return render(request, 'expenses/predictions.html', context)


@login_required
def recurring_payments(request):
    """Seznam trvalých plateb"""
    today = timezone.now().date()
    next_30_days = today + timedelta(days=30)
    
    payments = RecurringPayment.objects.all().order_by('next_payment_date')
    
    # Rozdělit na nadcházející (next 30 days) a potvrzené (již vytvořené)
    # Zobrazit VŠECHNY platby - historické i budoucí
    upcoming_payments = []
    confirmed_payments = []
    historical_payments = []
    
    for payment in payments:
        # Najít VŠECHNY transakce pro tuto trvalou platbu
        # 1. Transakce přímo propojené přes recurring_payment
        linked_transactions = Transaction.objects.filter(
            recurring_payment=payment
        ).order_by('-date')
        
        # 2. Transakce, které odpovídají podle detailů (kategorie, subkategorie, částka, typ)
        # ale nejsou propojené přes recurring_payment (pro případ, že byly vytvořeny ručně)
        matching_transactions = Transaction.objects.filter(
            category=payment.category,
            subcategory=payment.subcategory,
            amount=payment.amount,
            transaction_type=payment.transaction_type
        ).filter(
            Q(recurring_payment__isnull=True) | Q(recurring_payment=payment)
        ).exclude(
            id__in=linked_transactions.values_list('id', flat=True)
        ).order_by('-date')
        
        # Kombinovat a seřadit podle data, odstranit duplikáty podle ID
        all_confirmed_transactions = list(linked_transactions)
        seen_ids = {t.id for t in all_confirmed_transactions}
        for t in matching_transactions:
            if t.id not in seen_ids:
                all_confirmed_transactions.append(t)
                seen_ids.add(t.id)
        
        all_confirmed_transactions.sort(key=lambda x: x.date, reverse=True)
        
        # Zkontrolovat, zda existuje transakce pro aktuální next_payment_date
        has_current_transaction = Transaction.objects.filter(
            Q(recurring_payment=payment, date=payment.next_payment_date) |
            Q(category=payment.category, subcategory=payment.subcategory, 
              amount=payment.amount, date=payment.next_payment_date,
              transaction_type=payment.transaction_type)
        ).exists()
        
        is_upcoming = payment.next_payment_date <= next_30_days and payment.next_payment_date >= today
        is_past = payment.next_payment_date < today
        
        # Přidat všechny potvrzené transakce do seznamu
        for transaction in all_confirmed_transactions:
            confirmed_payments.append({
                'payment': payment,
                'transaction': transaction,
                'transaction_date': transaction.date,
                'month': transaction.date.strftime('%Y-%m')
            })
        
        # Pokud není potvrzená a je v příštích 30 dnech, přidat do nadcházejících
        if not has_current_transaction and is_upcoming:
            upcoming_payments.append({
                'payment': payment,
                'has_transaction': False,
                'month': payment.next_payment_date.strftime('%Y-%m')
            })
        # Pokud není potvrzená a je v minulosti, přidat do historických
        elif not has_current_transaction and is_past:
            historical_payments.append({
                'payment': payment,
                'has_transaction': False,
                'month': payment.next_payment_date.strftime('%Y-%m')
            })
    
    if request.method == 'POST':
        form = RecurringPaymentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Trvalá platba byla přidána.')
            return redirect('recurring_payments')
    else:
        form = RecurringPaymentForm()
    
    return render(request, 'expenses/recurring_payments.html', {
        'upcoming_payments': upcoming_payments,
        'confirmed_payments': confirmed_payments,
        'historical_payments': historical_payments,
        'form': form
    })


@login_required
def create_transaction_from_recurring(request, pk):
    """Vytvoření transakce z trvalé platby"""
    recurring_payment = get_object_or_404(RecurringPayment, pk=pk)
    
    # Zkontrolovat, zda už existuje transakce pro tuto trvalou platbu a datum
    existing_by_recurring = Transaction.objects.filter(
        recurring_payment=recurring_payment,
        date=recurring_payment.next_payment_date
    ).first()
    
    # Zkontrolovat také podle category, subcategory, amount a date (pro duplikáty)
    # Toto je hlavní kontrola - pokud existuje transakce s těmito údaji, nelze vytvořit novou
    existing_by_details = Transaction.objects.filter(
        category=recurring_payment.category,
        subcategory=recurring_payment.subcategory,
        amount=recurring_payment.amount,
        date=recurring_payment.next_payment_date,
        transaction_type=recurring_payment.transaction_type
    ).first()
    
    # Použít první nalezenou transakci
    existing_transaction = existing_by_recurring or existing_by_details
    
    # Pokud existuje duplikát podle detailů (ne jen podle recurring_payment), nelze vytvořit
    if existing_by_details and not existing_by_recurring:
        messages.warning(request, f'Transakce s těmito údaji (kategorie, subkategorie, částka, datum) již existuje a nemůže být vytvořena znovu.')
        return redirect('recurring_payments')
    
    if request.method == 'POST':
        if existing_transaction:
            # Pokud existuje a uživatel potvrdil přepsání
            if request.POST.get('overwrite') == 'yes':
                # Aktualizovat existující transakci
                existing_transaction.amount = recurring_payment.amount
                existing_transaction.description = recurring_payment.name
                existing_transaction.transaction_type = recurring_payment.transaction_type
                existing_transaction.category = recurring_payment.category
                existing_transaction.subcategory = recurring_payment.subcategory
                existing_transaction.payment_for = recurring_payment.payment_for
                existing_transaction.note = recurring_payment.note
                existing_transaction.save()
                messages.success(request, f'Transakce "{recurring_payment.name}" byla aktualizována.')
            else:
                messages.info(request, 'Transakce nebyla vytvořena.')
                return redirect('recurring_payments')
        else:
            # Zkontrolovat znovu před vytvořením (pro případ, že by někdo mezitím vytvořil)
            duplicate_check = Transaction.objects.filter(
                category=recurring_payment.category,
                subcategory=recurring_payment.subcategory,
                amount=recurring_payment.amount,
                date=recurring_payment.next_payment_date,
                transaction_type=recurring_payment.transaction_type
            ).exists()
            
            if duplicate_check:
                messages.warning(request, f'Transakce s těmito údaji již existuje a nemůže být vytvořena znovu.')
                return redirect('recurring_payments')
            
            # Vytvořit novou transakci
            transaction = Transaction.objects.create(
                amount=recurring_payment.amount,
                description=recurring_payment.name,
                transaction_type=recurring_payment.transaction_type,
                category=recurring_payment.category,
                subcategory=recurring_payment.subcategory,
                date=recurring_payment.next_payment_date,
                payment_for=recurring_payment.payment_for,
                note=recurring_payment.note,
                created_by=request.user,
                recurring_payment=recurring_payment,
                approved=False
            )
            messages.success(request, f'Transakce "{recurring_payment.name}" byla vytvořena.')
        
        # Aktualizovat datum další platby - přidat měsíce pouze pokud bude v příštích 30 dnech
        today = timezone.now().date()
        next_30_days = today + timedelta(days=30)
        
        payment_date = recurring_payment.next_payment_date
        year = payment_date.year
        month = payment_date.month + recurring_payment.frequency_months
        day = payment_date.day
        
        # Zpracovat přetečení měsíců
        while month > 12:
            month -= 12
            year += 1
        
        # Zkontrolovat, zda den existuje v novém měsíci (např. 31. ledna -> 31. února neexistuje)
        from calendar import monthrange
        max_day = monthrange(year, month)[1]
        if day > max_day:
            day = max_day
        
        from datetime import date
        new_payment_date = date(year, month, day)
        
        # Aktualizovat pouze pokud nové datum je v příštích 30 dnech
        if new_payment_date <= next_30_days:
            recurring_payment.next_payment_date = new_payment_date
            recurring_payment.save()
        # Pokud nové datum je mimo 30 dní, ponechat původní datum (nebude se zobrazovat v seznamu)
        
        return redirect('recurring_payments')
    
    # GET request - zobrazit potvrzení
    context = {
        'recurring_payment': recurring_payment,
        'existing_transaction': existing_transaction,
    }
    return render(request, 'expenses/confirm_recurring_transaction.html', context)


@login_required
def investments(request):
    """Přehled investičních skupin"""
    investments_list = Investment.objects.select_related('owner').all().order_by('-created_at')
    investment_rows = []
    sort_by = request.GET.get('sort', 'name')
    sort_order = request.GET.get('order', 'asc')
    if sort_order not in ['asc', 'desc']:
        sort_order = 'asc'

    # Vypočítat celkové hodnoty z nejnovějších pozorování
    total_invested = Decimal('0')
    total_current = Decimal('0')
    split_user1_label, split_user2_label = get_split_user_labels()
    total_invested_user1 = Decimal('0')
    total_invested_user2 = Decimal('0')
    total_current_user1 = Decimal('0')
    total_current_user2 = Decimal('0')

    for inv in investments_list:
        observations = list(inv.observations.all().order_by('-observation_date', '-created_at'))
        latest_observation = observations[0] if observations else None
        current_value = latest_observation.observed_value if latest_observation else (inv.observed_value or Decimal('0'))
        current_value_date = latest_observation.observation_date if latest_observation else inv.observed_value_date
        invested_amount = inv.invested_amount
        profit_loss = current_value - invested_amount if current_value else None
        profit_loss_percent = ((current_value - invested_amount) / invested_amount) * 100 if current_value and invested_amount else None

        total_invested += invested_amount
        total_current += current_value

        # Rozdělení statistik podle vlastníka investiční skupiny.
        # Fallback: pokud owner není vyplněný, použij uživatele, který zapsal
        # nejnovější investiční transakci v dané skupině.
        owner_username = (inv.owner.username if inv.owner else '').lower() if inv.owner else ''
        if not owner_username:
            owner_tx = inv.transactions.filter(
                transaction_type=TransactionType.INVESTMENT
            ).exclude(created_by__isnull=True).order_by('-created_at').first()
            owner_username = (owner_tx.created_by.username or '').lower() if owner_tx and owner_tx.created_by else ''

        if owner_username == split_user1_label.lower():
            total_invested_user1 += invested_amount
            total_current_user1 += current_value
        elif owner_username == split_user2_label.lower():
            total_invested_user2 += invested_amount
            total_current_user2 += current_value

        investment_rows.append({
            'investment': inv,
            'owner_name': inv.owner.username if inv.owner else '',
            'observations': observations,
            'latest_observation': latest_observation,
            'invested_amount': invested_amount,
            'observed_value': current_value if current_value else None,
            'observed_value_date': current_value_date,
            'profit_loss': profit_loss,
            'profit_loss_percent': profit_loss_percent,
            'has_more_observations': len(observations) > 1,
        })

    sort_key_map = {
        'name': lambda row: (row['investment'].name or '').lower(),
        'owner': lambda row: (row['owner_name'] or '').lower(),
        'invested': lambda row: row['invested_amount'] or Decimal('0'),
        'observed': lambda row: row['observed_value'] or Decimal('0'),
        'date': lambda row: row['observed_value_date'] or datetime.min.date(),
        'profit': lambda row: row['profit_loss'] or Decimal('0'),
        'percent': lambda row: row['profit_loss_percent'] or Decimal('0'),
    }
    if sort_by not in sort_key_map:
        sort_by = 'name'
    investment_rows = sorted(
        investment_rows,
        key=sort_key_map[sort_by],
        reverse=(sort_order == 'desc')
    )

    def owner_row_class_from_username(owner_user):
        if not owner_user:
            return ''
        name = (owner_user.username or '').lower()
        if name == 'zuzka':
            return 'owner-zuzka-row'
        if name == 'jirka':
            return 'owner-jirka-row'
        return ''

    for row in investment_rows:
        owner = row['investment'].owner
        row['owner_row_class'] = owner_row_class_from_username(owner)
        row['owner_badge_class'] = (
            'owner-badge-zuzka' if owner and owner.username.lower() == 'zuzka'
            else 'owner-badge-jirka' if owner and owner.username.lower() == 'jirka'
            else 'owner-badge-default'
        )

    total_profit_loss = total_current - total_invested
    total_profit_loss_user1 = total_current_user1 - total_invested_user1
    total_profit_loss_user2 = total_current_user2 - total_invested_user2
    
    context = {
        'investment_rows': investment_rows,
        'total_invested': total_invested,
        'total_current': total_current,
        'total_profit_loss': total_profit_loss,
        'total_invested_user1': total_invested_user1,
        'total_invested_user2': total_invested_user2,
        'total_current_user1': total_current_user1,
        'total_current_user2': total_current_user2,
        'total_profit_loss_user1': total_profit_loss_user1,
        'total_profit_loss_user2': total_profit_loss_user2,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'split_user1_label': split_user1_label,
        'split_user2_label': split_user2_label,
    }
    
    return render(request, 'expenses/investments.html', context)


@login_required
def edit_investment(request, pk):
    """Editace konkrétního záznamu pozorování + přidání nového pozorování."""
    investment = get_object_or_404(Investment, pk=pk)
    latest_observation = investment.observations.order_by('-observation_date', '-created_at').first()
    observation_id = request.GET.get('observation_id') or request.POST.get('observation_id')
    if observation_id:
        observation_to_edit = get_object_or_404(InvestmentObservation, pk=observation_id, investment=investment)
    else:
        observation_to_edit = latest_observation

    def sync_latest_observation_fields(target_investment):
        latest_observation = target_investment.observations.order_by('-observation_date', '-created_at').first()
        if latest_observation:
            target_investment.observed_value = latest_observation.observed_value
            target_investment.observed_value_date = latest_observation.observation_date
        else:
            target_investment.observed_value = None
            target_investment.observed_value_date = None
        target_investment.save(update_fields=['observed_value', 'observed_value_date', 'updated_at'])

    edit_initial = {}
    if observation_to_edit:
        edit_initial['observation_date'] = observation_to_edit.observation_date
    edit_form = InvestmentValueForm(instance=observation_to_edit, initial=edit_initial)
    add_form = InvestmentValueForm(initial={'observation_date': timezone.now().date()})

    if request.method == 'POST':
        if 'update_observation' in request.POST:
            if not observation_to_edit:
                messages.error(request, 'Není vybraný záznam k úpravě.')
                return redirect('edit_investment', pk=investment.pk)
            edit_form = InvestmentValueForm(
                request.POST,
                instance=observation_to_edit,
                initial={'observation_date': observation_to_edit.observation_date}
            )
            if edit_form.is_valid():
                edit_form.save()
                sync_latest_observation_fields(investment)
                messages.success(request, 'Záznam pozorování byl upraven.')
                return redirect(f"{reverse('edit_investment', kwargs={'pk': investment.pk})}?observation_id={observation_to_edit.pk}")
        elif 'add_observation' in request.POST:
            add_form = InvestmentValueForm(request.POST, initial={'observation_date': timezone.now().date()})
            if add_form.is_valid():
                new_observation = add_form.save(commit=False)
                new_observation.investment = investment
                new_observation.save()
                sync_latest_observation_fields(investment)
                messages.success(request, 'Nové pozorování bylo přidáno.')
                return redirect(f"{reverse('edit_investment', kwargs={'pk': investment.pk})}?observation_id={new_observation.pk}")

    return render(request, 'expenses/edit_investment.html', {
        'edit_form': edit_form,
        'add_form': add_form,
        'investment': investment,
        'observation_to_edit': observation_to_edit,
    })


@login_required
def settings(request):
    """Nastavení - kategorie a subkategorie"""
    categories = Category.objects.prefetch_related('subcategories').all()
    subcategories = Subcategory.objects.select_related('category').all()
    category_form = CategoryForm()
    subcategory_form = SubcategoryForm()
    investment_form = InvestmentForm(initial={'owner': request.user})
    
    # Hledání anomálií
    anomalies = []
    
    # Příliš vysoké položky (nad 3x průměr kategorie)
    for category in categories:
        category_transactions = Transaction.objects.filter(category=category, approved=True)
        if category_transactions.exists():
            avg_amount = category_transactions.aggregate(avg=Sum('amount'))['avg'] / category_transactions.count()
            high_transactions = category_transactions.filter(amount__gt=avg_amount * 3)
            for trans in high_transactions:
                anomalies.append({
                    'type': 'high_amount',
                    'transaction': trans,
                    'category': category,
                    'avg': avg_amount,
                })
    
    # Neodpovídající částky pro kategorii (statisticky odlehlé hodnoty)
    for category in categories:
        category_transactions = Transaction.objects.filter(category=category, approved=True)
        if category_transactions.count() > 5:
            amounts = [float(t.amount) for t in category_transactions]
            if amounts:
                mean = statistics.mean(amounts)
                stdev = statistics.stdev(amounts) if len(amounts) > 1 else 0
                if stdev > 0:
                    for trans in category_transactions:
                        z_score = abs((float(trans.amount) - mean) / stdev) if stdev > 0 else 0
                        if z_score > 3:  # 3 sigma rule
                            anomalies.append({
                                'type': 'outlier',
                                'transaction': trans,
                                'category': category,
                                'z_score': z_score,
                            })
    
    if request.method == 'POST':
        if 'add_category' in request.POST:
            category_form = CategoryForm(request.POST)
            if category_form.is_valid():
                category_form.save()
                messages.success(request, 'Kategorie byla přidána.')
                return redirect('settings')
        elif 'add_subcategory' in request.POST:
            subcategory_form = SubcategoryForm(request.POST)
            if subcategory_form.is_valid():
                subcategory_form.save()
                messages.success(request, 'Subkategorie byla přidána.')
                return redirect('settings')
        elif 'add_investment' in request.POST:
            investment_form = InvestmentForm(request.POST)
            if investment_form.is_valid():
                investment = investment_form.save(commit=False)
                if not investment.owner:
                    investment.owner = request.user
                investment.save()
                messages.success(request, 'Investiční skupina byla přidána.')
                return redirect('settings')
    
    context = {
        'categories': categories,
        'subcategories': subcategories,
        'category_form': category_form,
        'subcategory_form': subcategory_form,
        'investment_form': investment_form,
        'anomalies': anomalies[:20],  # Limit na 20 anomálií
    }
    
    return render(request, 'expenses/settings.html', context)


@login_required
def get_subcategories(request):
    """AJAX endpoint pro načtení subkategorií podle kategorie"""
    category_id = request.GET.get('category_id')
    if category_id:
        subcategories = Subcategory.objects.filter(category_id=category_id).values('id', 'name')
        return JsonResponse(list(subcategories), safe=False)
    return JsonResponse([], safe=False)


@login_required
def export_transactions(request):
    """Export transakcí do CSV nebo Excel"""
    # Get date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    export_format = request.GET.get('format', 'csv')  # 'csv' or 'excel'
    
    # Get all transactions (exclude deleted)
    transactions = Transaction.objects.filter(is_deleted=False)
    
    # Apply date filters
    if date_from:
        transactions = transactions.filter(date__gte=date_from)
    if date_to:
        transactions = transactions.filter(date__lte=date_to)
    
    # Order by date
    transactions = transactions.order_by('date', 'created_at')
    
    if export_format == 'excel' and OPENPYXL_AVAILABLE:
        # Export to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Transakce"
        
        # Headers
        headers = [
            'Datum', 'Popis', 'Typ', 'Kategorie', 'Subkategorie', 
            'Částka (Kč)', 'Za koho', 'Na kolik měsíců', 
            'Schváleno', 'Poznámka', 'Kdo zapsal', 'Datum zapsání', 'Importováno'
        ]
        ws.append(headers)
        
        # Data
        for t in transactions:
            ws.append([
                t.date.strftime('%Y-%m-%d') if t.date else '',
                t.description,
                t.get_transaction_type_display(),
                t.category.name if t.category else '',
                t.subcategory.name if t.subcategory else '',
                float(t.amount),
                t.get_payment_for_display(),
                t.months_duration,
                'Ano' if t.approved else 'Ne',
                t.note,
                t.created_by.username if t.created_by else '',
                t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else '',
                'Ano' if t.is_imported else 'Ne',
            ])
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'transakce_{date_from or "all"}_{date_to or "all"}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    else:
        # Export to CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'transakce_{date_from or "all"}_{date_to or "all"}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Add BOM for Excel UTF-8 compatibility
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'Datum', 'Popis', 'Typ', 'Kategorie', 'Subkategorie', 
            'Částka (Kč)', 'Za koho', 'Na kolik měsíců', 
            'Schváleno', 'Poznámka', 'Kdo zapsal', 'Datum zapsání', 'Importováno'
        ])
        
        # Data
        for t in transactions:
            writer.writerow([
                t.date.strftime('%Y-%m-%d') if t.date else '',
                t.description,
                t.get_transaction_type_display(),
                t.category.name if t.category else '',
                t.subcategory.name if t.subcategory else '',
                float(t.amount),
                t.get_payment_for_display(),
                t.months_duration,
                'Ano' if t.approved else 'Ne',
                t.note,
                t.created_by.username if t.created_by else '',
                t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else '',
                'Ano' if t.is_imported else 'Ne',
            ])
        
        return response


@login_required
def import_transactions(request):
    """Import transakcí z Excel souboru"""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Check file extension
        if not (file.name.endswith('.xlsx') or file.name.endswith('.xls') or file.name.endswith('.csv')):
            messages.error(request, 'Nepodporovaný formát souboru. Použijte .xlsx, .xls nebo .csv.')
            return redirect('manage_transactions?tab=import-export')
        
        try:
            imported_count = 0
            errors = []
            
            if file.name.endswith('.csv'):
                # Parse CSV
                file_content = file.read().decode('utf-8-sig')  # Handle BOM
                csv_reader = csv.DictReader(io.StringIO(file_content))
                
                # Expected headers (flexible matching)
                header_mapping = {
                    'datum': 'date',
                    'popis': 'description',
                    'typ': 'transaction_type',
                    'kategorie': 'category',
                    'subkategorie': 'subcategory',
                    'částka (kč)': 'amount',
                    'za koho': 'payment_for',
                    'na kolik měsíců': 'months_duration',
                    'schváleno': 'approved',
                    'poznámka': 'note',
                }
                
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        # Map headers (case-insensitive)
                        row_lower = {k.lower().strip(): v for k, v in row.items()}
                        
                        # Extract data
                        date_str = row_lower.get('datum', '').strip()
                        description = row_lower.get('popis', '').strip()
                        type_str = row_lower.get('typ', '').strip()
                        category_name = row_lower.get('kategorie', '').strip()
                        subcategory_name = row_lower.get('subkategorie', '').strip()
                        amount_str = row_lower.get('částka (kč)', '').strip()
                        
                        if not date_str or not description or not amount_str:
                            errors.append(f'Řádek {row_num}: Chybí povinná pole')
                            continue
                        
                        # Parse date
                        try:
                            date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except:
                            errors.append(f'Řádek {row_num}: Neplatné datum: {date_str}')
                            continue
                        
                        # Parse amount
                        try:
                            amount = Decimal(str(amount_str).replace(',', '.'))
                        except:
                            errors.append(f'Řádek {row_num}: Neplatná částka: {amount_str}')
                            continue
                        
                        # Map transaction type
                        type_map = {'příjem': TransactionType.INCOME, 'výdaj': TransactionType.EXPENSE, 
                                   'investice': TransactionType.INVESTMENT, 'investice (přesun)': TransactionType.INVESTMENT}
                        transaction_type = type_map.get(type_str.lower(), TransactionType.EXPENSE)
                        
                        # Find category
                        category = None
                        if category_name:
                            category = Category.objects.filter(name__iexact=category_name).first()
                        
                        # Find subcategory
                        subcategory = None
                        if subcategory_name and category:
                            subcategory = Subcategory.objects.filter(
                                category=category, 
                                name__iexact=subcategory_name
                            ).first()
                        
                        # Payment for
                        payment_for_str = row_lower.get('za koho', '').strip().lower()
                        payment_for_map = {'za sebe': PaymentFor.SELF, 'za partnera': PaymentFor.PARTNER, 
                                          'společný účet': PaymentFor.SHARED}
                        payment_for = payment_for_map.get(payment_for_str, PaymentFor.SELF)
                        
                        # Months duration
                        months_duration = 0
                        if row_lower.get('na kolik měsíců'):
                            try:
                                months_duration = int(row_lower.get('na kolik měsíců'))
                            except:
                                pass
                        
                        # Approved
                        approved_str = row_lower.get('schváleno', '').strip().lower()
                        approved = approved_str in ('ano', 'yes', 'true', '1')
                        
                        # Note
                        note = row_lower.get('poznámka', '').strip()
                        
                        # Create transaction
                        transaction = Transaction.objects.create(
                            date=date,
                            description=description,
                            transaction_type=transaction_type,
                            category=category,
                            subcategory=subcategory,
                            amount=amount,
                            payment_for=payment_for,
                            months_duration=months_duration,
                            approved=approved,
                            note=note,
                            created_by=request.user,
                            is_imported=True,
                            is_deleted=False
                        )
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f'Řádek {row_num}: Chyba - {str(e)}')
            
            elif OPENPYXL_AVAILABLE:
                # Parse Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                # Read headers
                headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
                
                # Process rows
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        
                        # Extract data (same logic as CSV)
                        date_str = str(row_dict.get('datum', '')).strip()
                        description = str(row_dict.get('popis', '')).strip()
                        type_str = str(row_dict.get('typ', '')).strip()
                        category_name = str(row_dict.get('kategorie', '')).strip()
                        subcategory_name = str(row_dict.get('subkategorie', '')).strip()
                        amount_str = str(row_dict.get('částka (kč)', '')).strip()
                        
                        if not date_str or not description or not amount_str:
                            continue
                        
                        # Parse date
                        if isinstance(row_dict.get('datum'), datetime):
                            date = row_dict['datum'].date()
                        else:
                            try:
                                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                            except:
                                errors.append(f'Řádek {row_num}: Neplatné datum: {date_str}')
                                continue
                        
                        # Parse amount
                        try:
                            if isinstance(row_dict.get('částka (kč)'), (int, float)):
                                amount = Decimal(str(row_dict['částka (kč)']))
                            else:
                                amount = Decimal(str(amount_str).replace(',', '.'))
                        except:
                            errors.append(f'Řádek {row_num}: Neplatná částka: {amount_str}')
                            continue
                        
                        # Map transaction type
                        type_map = {'příjem': TransactionType.INCOME, 'výdaj': TransactionType.EXPENSE, 
                                   'investice': TransactionType.INVESTMENT, 'investice (přesun)': TransactionType.INVESTMENT}
                        transaction_type = type_map.get(type_str.lower(), TransactionType.EXPENSE)
                        
                        # Find category
                        category = None
                        if category_name:
                            category = Category.objects.filter(name__iexact=category_name).first()
                        
                        # Find subcategory
                        subcategory = None
                        if subcategory_name and category:
                            subcategory = Subcategory.objects.filter(
                                category=category, 
                                name__iexact=subcategory_name
                            ).first()
                        
                        # Payment for
                        payment_for_str = str(row_dict.get('za koho', '')).strip().lower()
                        payment_for_map = {'za sebe': PaymentFor.SELF, 'za partnera': PaymentFor.PARTNER, 
                                          'společný účet': PaymentFor.SHARED}
                        payment_for = payment_for_map.get(payment_for_str, PaymentFor.SELF)
                        
                        # Months duration
                        months_duration = 0
                        if row_dict.get('na kolik měsíců'):
                            try:
                                months_duration = int(row_dict.get('na kolik měsíců'))
                            except:
                                pass
                        
                        # Approved
                        approved_str = str(row_dict.get('schváleno', '')).strip().lower()
                        approved = approved_str in ('ano', 'yes', 'true', '1')
                        
                        # Note
                        note = str(row_dict.get('poznámka', '')).strip()
                        
                        # Create transaction
                        transaction = Transaction.objects.create(
                            date=date,
                            description=description,
                            transaction_type=transaction_type,
                            category=category,
                            subcategory=subcategory,
                            amount=amount,
                            payment_for=payment_for,
                            months_duration=months_duration,
                            approved=approved,
                            note=note,
                            created_by=request.user,
                            is_imported=True,
                            is_deleted=False
                        )
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f'Řádek {row_num}: Chyba - {str(e)}')
            else:
                messages.error(request, 'Pro import Excel souborů je potřeba nainstalovat openpyxl: pip install openpyxl')
                return redirect('manage_transactions?tab=import-export')
            
            if imported_count > 0:
                messages.success(request, f'Úspěšně importováno {imported_count} transakcí.')
            if errors:
                messages.warning(request, f'Při importu došlo k {len(errors)} chybám. Zkontrolujte data.')
            
            return redirect('manage_transactions?tab=manage')
            
        except Exception as e:
            messages.error(request, f'Chyba při importu: {str(e)}')
            return redirect('manage_transactions?tab=import-export')
    
    return redirect('manage_transactions?tab=import-export')


@login_required
def remove_transaction(request, pk):
    """Soft delete transakce (pouze pro importované)"""
    transaction = get_object_or_404(Transaction, pk=pk)
    
    if not transaction.is_imported:
        messages.error(request, 'Lze odstranit pouze importované transakce.')
        return redirect('manage_transactions?tab=manage')
    
    if request.method == 'POST':
        transaction.is_deleted = True
        transaction.save()
        messages.success(request, 'Transakce byla odstraněna.')
        return redirect('manage_transactions?tab=manage')
    
    # GET request - show confirmation
    return render(request, 'expenses/confirm_remove_transaction.html', {'transaction': transaction})

